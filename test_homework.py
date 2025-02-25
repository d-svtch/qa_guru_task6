from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from conftest import ZIP_PATH


def test_pdf_reader():
    with ZipFile(ZIP_PATH, "r") as zip_file:
        with zip_file.open("files/test_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            assert "This is pdf file for tests" == reader.pages[0].extract_text()


def test_xlsx_reader():
    with ZipFile(ZIP_PATH, "r") as zip_file:
        with zip_file.open("files/test_xlsx.xlsx") as xlsx_file:
            reader = load_workbook(xlsx_file)
            assert "TEST" == reader.active.cell(row=3, column=2).value


def test_csv_reader():
    with ZipFile(ZIP_PATH, "r") as zip_file:
        with zip_file.open("files/test_csv.csv") as csv_file:
            reader = csv_file.read().decode("utf-8")
            assert "Jenkins" in reader
